from __future__ import annotations

import csv
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / "data" / "sample.xlsx"
OUTPUT = ROOT / "data" / "sample_alerts.csv"


def clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.isoformat()
    return str(value).replace("\r\n", "\n").strip()


def main() -> None:
    workbook = load_workbook(INPUT, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = sheet.iter_rows(values_only=True)
    headers = [clean(value) for value in next(rows)]
    indexes = {header: index for index, header in enumerate(headers) if header}

    output_by_alert_id = {}
    for row in rows:
        alert_id = clean(row[indexes["alert_id"]])
        if not alert_id:
            continue

        remarks = clean(row[indexes["L2_remarks"]])
        if not remarks:
            remarks = clean(row[indexes["L1_remarks"]])

        output_by_alert_id[alert_id] = {
            "alert_id": alert_id,
            "analyst": clean(row[indexes["L2_agent"]]) or clean(row[indexes["L1_agent"]]),
            "l1_agent": clean(row[indexes["L1_agent"]]),
            "l1_remarks": clean(row[indexes["L1_remarks"]]),
            "l2_agent": clean(row[indexes["L2_agent"]]),
            "l2_remarks": clean(row[indexes["L2_remarks"]]),
            "investigation_remarks": remarks,
            "updated_at": clean(row[indexes["L2_decision_date"]])
            or clean(row[indexes["L1_decision_date"]]),
            "case_status": clean(row[indexes["L2_decision"]]) or clean(row[indexes["L1_decision"]]),
            "risk_level": clean(row[indexes["risk_rating"]]),
        }

    output_rows = list(output_by_alert_id.values())

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with OUTPUT.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "alert_id",
                "analyst",
                "l1_agent",
                "l1_remarks",
                "l2_agent",
                "l2_remarks",
                "investigation_remarks",
                "updated_at",
                "case_status",
                "risk_level",
            ],
        )
        writer.writeheader()
        writer.writerows(output_rows)

    print(f"Wrote {len(output_rows)} unique alert rows to {OUTPUT}")


if __name__ == "__main__":
    main()
